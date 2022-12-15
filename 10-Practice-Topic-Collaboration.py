import tabula
import os
import pandas as pd
from datetime import date
today = date.today().strftime("%Y%m%d") ## กำหนดวันที่

## Step 1 #Extract PDF to CSV to XLS ######################
path = 'C:/Users/KrOm/Documents/aritnet/Practice-Topic/' #กำหนด folder ที่จะให้ระบบทำงาน
os.chdir(path) #รัน Code เพื่อสั่งให้ระบบทำงานบน folder ที่ต้องการ
file1 = today+".pdf"
table = tabula.read_pdf(file1,pages=1)
table[0]
tabula.convert_into(today+".pdf",today+".csv")
read_file = pd.read_csv (today+'.csv')
read_file.to_excel (today+'.xlsx', index = None, header=True)

## Step 2 #Automate MS-Excel ######################
import openpyxl as xl
from openpyxl.chart import LineChart, Reference
import win32com.client
from PIL import ImageGrab
    ######## Generate automated excel workbook ########
workbook = xl.load_workbook(today+'.xlsx')
sheet_1 = workbook['Sheet1']
for row in range(2, sheet_1.max_row + 1):
    current = sheet_1.cell(row, 2)
    voltage = sheet_1.cell(row, 3)
    power = float(current.value) * float(voltage.value)
    power_cell = sheet_1.cell(row, 1)
    power_cell.value = power
values = Reference(sheet_1, min_row = 2, max_row = sheet_1.max_row, min_col = 1, max_col = 1)
chart = LineChart()
chart.y_axis.title = 'Power'
chart.x_axis.title = 'Index'
chart.add_data(values)
sheet_1.add_chart(chart, 'e2')
workbook.save(today+'.xlsx')
    ######## Extract chart image from Excel workbook ########
input_file = "C:/Users/KrOm/Documents/aritnet/Practice-Topic/"+today+".xlsx"
output_image = "C:/Users/KrOm/Documents/aritnet/Practice-Topic/"+today+".png"
operation = win32com.client.Dispatch("Excel.Application")
operation.Visible = 0
operation.DisplayAlerts = 0 
workbook_2 = operation.Workbooks.Open(input_file)
sheet_2 = operation.Sheets(1)  
for x, chart in enumerate(sheet_2.Shapes):
    chart.Copy()
    image = ImageGrab.grabclipboard()
    image.save(output_image, 'png')
    pass
workbook_2.Close(True)
operation.Quit()

## Step 3 Flip Image ############################################
from PIL import Image 
# Flip an Image
flip_image = Image.open(today+'.png') 
flip_image = flip_image.transpose (Image.FLIP_LEFT_RIGHT)
flip_image.save(today+'-flip.png')

## Step 4 Image to PDF ###########################################
from PIL import Image
def Images_Pdf(filename, output):
    images = []
    for file in filename:
        im = Image.open(file)
        im = im.convert('RGB')
        images.append(im)
    images[0].save(output, save_all=True, append_images=images[1:])
Images_Pdf([today+"-flip.png"], today+"-flip.pdf")

## Step 5 Locked File CSV ########################################
from cryptography.fernet import Fernet
key = Fernet.generate_key()
with open(today+'.key', 'wb') as filekey:
   filekey.write(key)
with open(today+'.key', 'rb') as filekey:
    key = filekey.read()
# opening the original file to encrypt
with open(today+'.csv', 'rb') as file:
    original = file.read() 
# encrypting the file
f=Fernet(key)
encrypted = f.encrypt(original)
# opening the file in write mode and
# writing the encrypted data
with open(today+'_en.csv', 'wb') as encrypted_file:
    encrypted_file.write(encrypted)

## Step 6 Convert Mp4 to Mp3 and Split Audio 10 Sec ##############
from pydub import AudioSegment
from pydub.utils import mediainfo
from pydub.playback import play
# Extract Sound from Video
sound=AudioSegment.from_file(today+".mp4", format="mp4") 
sound.export(today+".mp3", format="mp3")
# Split Audio
sound = AudioSegment.from_file(today+".mp3", format="mp3")
sound_1=sound[:10000]
sound_1.export(today+".mp3", format="mp3") 

## Step 7 Proof Reading from Log File ##########################
#open text file in read mode
text_log = open(today+".log", "r")
raw_log = text_log.read()
text_log.close()

from textblob import TextBlob
sentence = TextBlob(raw_log)
contentmail = sentence.correct()

## Step 8 Send Email ###########################################
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.header import Header
import smtplib

user = 'krommavut.n@rmutsb.ac.th'
app_password = 'kvlklqvhydbusixv'
host = 'smtp.gmail.com'
port = 465
to = 'krommavut@gmail.com'
cc='nattapong.s@rmutsb.ac.th'

subject = 'AritNet Report '+today
content_txt = " "+str(contentmail)
attachment = '20221215-flip.pdf'
attachment2 = '20221215-en.csv'
attachment3 = '20221215.key'
attachment4 = '20221215.mp3'

### Define email ###
message = MIMEMultipart()
# add From 
message['From'] = Header(user)
# add To
message['To'] = Header(to)     
message['Cc'] = Header(cc) 
# add Subject
message['Subject'] = Header(subject)
# add content text
message.attach(MIMEText(content_txt, 'plain', 'utf-8'))
# add attachment
att_name = os.path.basename(attachment)
att1 = MIMEText(open(attachment, 'rb').read(), 'base64', 'utf-8')
att1['Content-Type'] = 'application/octet-stream'
att1['Content-Disposition'] = 'attachment; filename=' + att_name
message.attach(att1)

att_name = os.path.basename(attachment2)
att1 = MIMEText(open(attachment, 'rb').read(), 'base64', 'utf-8')
att1['Content-Type'] = 'application/octet-stream'
att1['Content-Disposition'] = 'attachment; filename=' + att_name
message.attach(att1)

att_name = os.path.basename(attachment3)
att1 = MIMEText(open(attachment, 'rb').read(), 'base64', 'utf-8')
att1['Content-Type'] = 'application/octet-stream'
att1['Content-Disposition'] = 'attachment; filename=' + att_name
message.attach(att1)

att_name = os.path.basename(attachment4)
att1 = MIMEText(open(attachment, 'rb').read(), 'base64', 'utf-8')
att1['Content-Type'] = 'application/octet-stream'
att1['Content-Disposition'] = 'attachment; filename=' + att_name
message.attach(att1)
    
### Send email ###
server = smtplib.SMTP_SSL(host, port) 
server.login(user, app_password)
server.sendmail(user, to, message.as_string()) 
server.quit() 
print('Sent email successfully')  